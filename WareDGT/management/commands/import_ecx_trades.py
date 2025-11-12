from decimal import Decimal
import os
import re
from pathlib import Path
import difflib
import datetime

from django.core.files import File
from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth import get_user_model
from openpyxl import load_workbook

from django.utils import timezone
from WareDGT.models import (
    Warehouse,
    Commodity,
    EcxTrade,
    EcxTradeReceiptFile,
    SeedType,
    Company,
)


class Command(BaseCommand):
    help = "Import ECX trade records from an Excel sheet"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", help="Path to ECX Excel file")
        parser.add_argument(
            "--receipts-dir",
            help="Optional directory containing scanned receipts named after the "
                 "warehouse receipt number (e.g. 123456.jpg)",
        )
        parser.add_argument(
            "--user",
            help="Username to assign as the creator of the records. Defaults to the first user.",
        )

    def handle(self, *args, **options):
        xlsx_path = options["xlsx_path"]
        receipts_dir = options.get("receipts_dir")
        username = options.get("user")

        if not os.path.exists(xlsx_path):
            raise CommandError(f"File not found: {xlsx_path}")

        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        def normalize_letters(name: str) -> str:
            return "".join(c for c in name.upper() if c.isalpha())

        # Match warehouses regardless of type to support DGT or ECX imports
        warehouses = {normalize_letters(w.name): w for w in Warehouse.objects.all()}

        def longest_common_substring(a: str, b: str) -> int:
            """Return length of longest common substring."""
            m = [[0] * (len(b) + 1) for _ in range(len(a) + 1)]
            longest = 0
            for i, ca in enumerate(a, 1):
                for j, cb in enumerate(b, 1):
                    if ca == cb:
                        m[i][j] = m[i - 1][j - 1] + 1
                        if m[i][j] > longest:
                            longest = m[i][j]
            return longest

        def match_warehouse(name: str):
            key = normalize_letters(name)
            wh = warehouses.get(key)
            if wh:
                return wh
            best = None
            best_len = 0
            for db_key, w in warehouses.items():
                lcs = longest_common_substring(db_key, key)
                if lcs >= 3 and lcs > best_len:
                    best_len = lcs
                    best = w
            return best
        def normalize(name: str) -> str:
            return re.sub(r"[^A-Z0-9]", "", name.upper())

        # Normalized lookup table using all warehouses without restricting by type
        warehouses = {normalize(w.name): w for w in Warehouse.objects.all()}
        current_wh = None
        header_seen = False
        last_item = None

        User = get_user_model()
        if username:
            try:
                user = User.objects.get(username=username)
            except User.DoesNotExist:
                raise CommandError(f"User '{username}' not found")
        else:
            user = User.objects.first()
            if not user:
                raise CommandError("No users exist to assign created_by")

        owner = Company.objects.filter(name="DGT").first()

        for row in ws.iter_rows(values_only=True):
            if not any(row):
                continue
            cell0 = str(row[0]).strip() if row[0] is not None else ""

            if cell0 and "WAREHOUSE" in cell0.upper():
                wh_name = re.sub("WAREHOUSE", "", cell0, flags=re.IGNORECASE).strip().upper()
                current_wh = match_warehouse(wh_name)
                wh_name = re.sub("WAREHOUSE", "", cell0, flags=re.IGNORECASE).strip()
                key = normalize(wh_name)
                current_wh = warehouses.get(key)
                if not current_wh:
                    match = difflib.get_close_matches(key, warehouses.keys(), n=1)
                    if match:
                        current_wh = warehouses[match[0]]
                if current_wh:
                    header_seen = False
                    self.stdout.write(self.style.SUCCESS(f"Processing {current_wh.name}"))
                else:
                    self.stdout.write(self.style.WARNING(f"Unknown warehouse '{wh_name}', skipping"))
                continue

            if not current_wh:
                continue

            if cell0.upper().startswith("PURCHASED ITEM TYPE"):
                header_seen = True
                continue

            if not header_seen:
                continue

            if isinstance(row[4], str) and row[4].startswith("="):
                # Skip formula summary rows
                continue

            item_type = str(row[0]).strip() if row[0] else last_item
            if not item_type:
                continue
            last_item = item_type

            net_receipt = row[1]
            wr_no = row[2]
            qty = row[3]
            purchase_date = row[9]

            if isinstance(purchase_date, datetime.datetime):
                purchase_date = purchase_date.date()
            elif isinstance(purchase_date, str):
                purchase_date_str = purchase_date.strip()
                parsed = None
                for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
                    try:
                        parsed = datetime.datetime.strptime(purchase_date_str, fmt).date()
                        break
                    except ValueError:
                        continue
                purchase_date = parsed

            if not all([net_receipt, wr_no, qty, purchase_date]):
                continue

            # split item code into symbol + grade if possible
            match = re.match(r"([A-Za-z]+)([0-9UG]*)", item_type)
            symbol = match.group(1) if match else item_type
            grade = match.group(2) if match else ""

            seed, _ = SeedType.objects.get_or_create(code=symbol, defaults={"name": symbol})
            commodity, _ = Commodity.objects.get_or_create(
                seed_type=seed,
                origin="",
                grade=grade,
            )

            status_val = row[11] if len(row) > 11 else None
            loaded = False
            if isinstance(status_val, str):
                status_clean = status_val.strip().lower()
                if status_clean in ["loaded", "delivery sent", "deliver sent"]:
                    loaded = True

            trade = EcxTrade.objects.create(
                warehouse=current_wh,
                commodity=commodity,
                net_obligation_receipt_no=str(net_receipt),
                warehouse_receipt_no=str(wr_no),
                quantity_quintals=Decimal(str(qty)),
                purchase_date=purchase_date,
                recorded_by=user,
                loaded=loaded,
                loaded_at=timezone.now() if loaded else None,
                owner=owner,
            )

            if receipts_dir:
                for ext in [".jpg", ".jpeg", ".png", ".pdf"]:
                    candidate = Path(receipts_dir) / f"{wr_no}{ext}"
                    if candidate.exists():
                        with open(candidate, "rb") as f:
                            EcxTradeReceiptFile.objects.create(
                                trade=trade,
                                file=File(f, name=candidate.name),
                            )
                        break

        self.stdout.write(self.style.SUCCESS("Import complete."))
