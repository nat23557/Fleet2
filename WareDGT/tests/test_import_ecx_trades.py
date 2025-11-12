from decimal import Decimal
import datetime
import os
import tempfile

import django
from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import TestCase

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "warehouse_project.settings_test")
django.setup()
call_command("migrate", verbosity=0)

from WareDGT.models import Warehouse, EcxTrade

from openpyxl import Workbook


class ImportEcxTradesCommandTests(TestCase):
    def _create_workbook(self):
        wb = Workbook()
        ws = wb.active
        # First warehouse
        ws.append(["SESAME", None, None, None, None, None, None, None, None, None])
        ws.append(
            [
                "ADDIS ABABA/SARIS WAREHOUSE",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        )
        ws.append(
            [
                "PURCHASED ITEM TYPE",
                "NET OBLIGATION RECEIPT NO.",
                "WAREHOUSE RECEIPT NO.",
                "QUINTAL",
                None,
                None,
                None,
                None,
                None,
                "PURCHASED DATE",
                "LAST PICK UP DATE",
                "Status",
            ]
        )
        ws.append(
            [
                "WWSSUG",
                "OR1",
                "WR1",
                10,
                None,
                None,
                None,
                None,
                None,
                datetime.date(2024, 1, 1),
                None,
                "Loaded",
            ]
        )
        # Formula summary row should be ignored
        ws.append(
            [
                None,
                None,
                None,
                None,
                "=SUM(D4)",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        )

        # Second warehouse with misspelled name for fuzzy match
        ws.append(["SESAME", None, None, None, None, None, None, None, None, None])
        ws.append(
            ["NEKAMT WAREHOUSE", None, None, None, None, None, None, None, None, None]
        )
        ws.append(
            [
                "PURCHASED ITEM TYPE",
                "NET OBLIGATION RECEIPT NO.",
                "WAREHOUSE RECEIPT NO.",
                "QUINTAL",
                None,
                None,
                None,
                None,
                None,
                "PURCHASED DATE",
                "LAST PICK UP DATE",
                "Status",
            ]
        )
        ws.append(
            [
                "WWSS4",
                "OR2",
                "WR2",
                20,
                None,
                None,
                None,
                None,
                None,
                datetime.date(2024, 2, 2),
                None,
                "Unloaded",
            ]
        )
        return wb

    def _create_stringdate_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["SESAME", None, None, None, None, None, None, None, None, None])
        ws.append(
            [
                "ADDIS ABABA/SARIS WAREHOUSE",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        )
        ws.append(
            [
                "PURCHASED ITEM TYPE",
                "NET OBLIGATION RECEIPT NO.",
                "WAREHOUSE RECEIPT NO.",
                "QUINTAL",
                None,
                None,
                None,
                None,
                None,
                "PURCHASED DATE",
                "LAST PICK UP DATE",
                "Status",
            ]
        )
        ws.append(
            [
                "WWSSUG",
                "OR5",
                "WR5",
                30,
                None,
                None,
                None,
                None,
                None,
                "11/13/2024",
                None,
                "Loaded",
            ]
        )
        return wb

    def _create_typo_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["SESAME", None, None, None, None, None, None, None, None, None])
        ws.append(
            [
                "ADDIS-ABABA-SARIS WAREHOUSE",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        )
        ws.append(
            [
                "PURCHASED ITEM TYPE",
                "NET OBLIGATION RECEIPT NO.",
                "WAREHOUSE RECEIPT NO.",
                "QUINTAL",
                None,
                None,
                None,
                None,
                None,
                "PURCHASED DATE",
                "LAST PICK UP DATE",
                "Status",
            ]
        )
        ws.append(
            [
                "WWSS1",
                "OR3",
                "WR3",
                15,
                None,
                None,
                None,
                None,
                None,
                datetime.date(2024, 3, 3),
                None,
                "Loaded",
            ]
        )

        ws.append(["SESAME", None, None, None, None, None, None, None, None, None])
        ws.append(
            ["Pawe Warehouse", None, None, None, None, None, None, None, None, None]
        )
        ws.append(
            [
                "PURCHASED ITEM TYPE",
                "NET OBLIGATION RECEIPT NO.",
                "WAREHOUSE RECEIPT NO.",
                "QUINTAL",
                None,
                None,
                None,
                None,
                None,
                "PURCHASED DATE",
                "LAST PICK UP DATE",
                "Status",
            ]
        )
        ws.append(
            [
                "WWSS2",
                "OR4",
                "WR4",
                25,
                None,
                None,
                None,
                None,
                None,
                datetime.date(2024, 4, 4),
                None,
                "Unloaded",
            ]
        )
        return wb

    def test_import_creates_movements_with_fuzzy_matching(self):
        User = get_user_model()
        user = User.objects.create_user(username="tester", password="pass")

        wh1 = Warehouse.objects.create(
            code="AA",
            name="ADDIS ABABA/SARIS",
            warehouse_type=Warehouse.ECX,
            capacity_quintals=Decimal("1000"),
            latitude=0,
            longitude=0,
        )
        wh2 = Warehouse.objects.create(
            code="NEK",
            name="NEKEMT",
            warehouse_type=Warehouse.ECX,
            capacity_quintals=Decimal("1000"),
            latitude=0,
            longitude=0,
        )

        wb = self._create_workbook()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name

        try:
            call_command("import_ecx_trades", tmp_path, "--user", user.username)
        finally:
            os.unlink(tmp_path)

        self.assertEqual(EcxTrade.objects.count(), 2)

        m1 = EcxTrade.objects.get(warehouse=wh1)
        self.assertEqual(m1.quantity_quintals, Decimal("10"))
        self.assertTrue(m1.loaded)
        self.assertIsNotNone(m1.loaded_at)

        m2 = EcxTrade.objects.get(warehouse=wh2)
        self.assertEqual(m2.quantity_quintals, Decimal("20"))
        self.assertFalse(m2.loaded)

    def test_import_handles_typo_and_punctuation_in_warehouse_names(self):
        User = get_user_model()
        user = User.objects.create_user(username="tester2", password="pass")

        wh1 = Warehouse.objects.create(
            code="AA",
            name="ADDIS ABABA/SARIS",
            warehouse_type=Warehouse.ECX,
            capacity_quintals=Decimal("1000"),
            latitude=0,
            longitude=0,
        )
        wh2 = Warehouse.objects.create(
            code="PAW",
            name="PAWE",
            warehouse_type=Warehouse.ECX,
            capacity_quintals=Decimal("1000"),
            latitude=0,
            longitude=0,
        )

        wb = self._create_typo_workbook()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name

        try:
            call_command("import_ecx_trades", tmp_path, "--user", user.username)
        finally:
            os.unlink(tmp_path)

        self.assertEqual(EcxTrade.objects.count(), 2)
        self.assertTrue(EcxTrade.objects.filter(warehouse=wh1).exists())
        self.assertTrue(EcxTrade.objects.filter(warehouse=wh2).exists())

    def test_import_parses_string_dates(self):
        User = get_user_model()
        user = User.objects.create_user(username="tester3", password="pass")

        wh = Warehouse.objects.create(
            code="AA",
            name="ADDIS ABABA/SARIS",
            warehouse_type=Warehouse.ECX,
            capacity_quintals=Decimal("1000"),
            latitude=0,
            longitude=0,
        )

        wb = self._create_stringdate_workbook()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name

        try:
            call_command("import_ecx_trades", tmp_path, "--user", user.username)
        finally:
            os.unlink(tmp_path)

        m = EcxTrade.objects.get(warehouse=wh)
        self.assertEqual(m.purchase_date, datetime.date(2024, 11, 13))
        self.assertTrue(m.loaded)
        self.assertIsNotNone(m.loaded_at)
