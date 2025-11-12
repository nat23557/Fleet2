from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError

from cash_management.models import BankAccount, Transaction


class Command(BaseCommand):
    help = "Import bank accounts and transactions from an Excel file (xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to Excel .xlsx file")
        parser.add_argument("--user", type=str, default=None, help="Username to attribute as creator")
        parser.add_argument("--sheet", type=str, default=None, help="Optional sheet name")
        parser.add_argument("--preview", action="store_true", help="Print header mapping and first rows; no writes")
        parser.add_argument("--default-account", dest="default_account", type=str, default=None, help="Fallback account name for rows missing account column")

    def handle(self, *args, **opts):
        path = Path(opts["file"]).expanduser()
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        try:
            import openpyxl  # type: ignore
        except Exception as exc:
            raise CommandError(
                "openpyxl not installed. Run: pip install openpyxl"
            ) from exc

        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb[opts["sheet"]] if opts.get("sheet") else wb.active

        # Expect a header row; normalize headers
        rows = ws.iter_rows(values_only=True)
        headers = next(rows)
        if not headers:
            raise CommandError("Empty sheet")
        header_map = {str(h).strip().lower(): idx for idx, h in enumerate(headers)}

        def pick(*names):
            for raw in names:
                for cand in (raw, raw.replace('_', ' ')):
                    idx = header_map.get(str(cand).strip().lower())
                    if idx is not None:
                        return idx
            return None

        # Candidate headers
        idx_account = pick('account', 'account_name', 'account name', 'bank account', 'bankaccount', 'account no', 'account number')
        idx_bank = pick('bank_name', 'bank name', 'bank')
        idx_currency = pick('currency', 'curr', 'ccy')
        idx_date = pick('date', 'transaction date', 'txn date', 'value date', 'posting date')
        idx_desc = pick('description', 'details', 'narration', 'particulars', 'remark', 'remarks', 'desc')
        idx_ref = pick('reference', 'ref', 'txn id', 'document no', 'voucher', 'cheque no', 'check no')
        idx_debit = pick('debit', 'dr', 'withdrawal', 'outflow', 'paid', 'payment', 'expense', 'debits')
        idx_credit = pick('credit', 'cr', 'deposit', 'inflow', 'received', 'income', 'credits')
        idx_amount = pick('amount', 'amt')
        idx_type = pick('type', 'dr_cr', 'direction', 'dc', 'd/c')

        if opts.get('preview'):
            self.stdout.write('Detected headers:')
            for k, v in header_map.items():
                self.stdout.write(f"  {v}: {k}")
            self.stdout.write('Column mapping:')
            self.stdout.write(f"  account={idx_account}, bank={idx_bank}, currency={idx_currency}, date={idx_date}")
            self.stdout.write(f"  desc={idx_desc}, ref={idx_ref}, debit={idx_debit}, credit={idx_credit}, amount={idx_amount}, type={idx_type}")
            # Show first 5 rows
            sample = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                sample.append(row)
                if len(sample) >= 5:
                    break
            for r in sample:
                self.stdout.write(f"  row: {list(r)}")
            return

        User = get_user_model()
        user = None
        if opts.get("user"):
            user = User.objects.filter(username=opts["user"]).first()

        created = 0
        for row in rows:
            if not row:
                continue
            account_name = (row[idx_account] if idx_account is not None else None) or opts.get('default_account')
            if not account_name:
                # Many ledgers carry account in title; allow default via flag
                continue
            bank_name = (row[idx_bank] if idx_bank is not None else '') or ''
            currency = (row[idx_currency] if idx_currency is not None else 'ETB') or 'ETB'
            acc, _ = BankAccount.objects.get_or_create(
                name=str(account_name).strip(),
                defaults={"bank_name": str(bank_name).strip(), "currency": str(currency).strip()},
            )
            if not acc.bank_name:
                acc.bank_name = str(bank_name).strip()
                acc.currency = str(currency).strip() or "ETB"
                acc.save(update_fields=["bank_name", "currency"])

            # Transaction
            dt = row[idx_date] if idx_date is not None else None
            desc = (row[idx_desc] if idx_desc is not None else '') or ''
            ref = (row[idx_ref] if idx_ref is not None else '') or ''
            debit = row[idx_debit] if idx_debit is not None else None
            credit = row[idx_credit] if idx_credit is not None else None
            amount = row[idx_amount] if idx_amount is not None else None
            ttype = (str(row[idx_type]).strip().lower() if (idx_type is not None and row[idx_type] is not None) else None)

            # If debit/credit not provided, infer from amount and type/sign
            if (debit in (None, '', 0)) and (credit in (None, '', 0)) and amount not in (None, ''):
                try:
                    amt = Decimal(str(amount).replace(',', ''))
                except Exception:
                    amt = Decimal(0)
                if ttype in ('dr', 'debit', 'out', 'outflow', 'withdrawal', 'payment', 'paid', 'expense'):
                    debit = amt
                    credit = Decimal(0)
                elif ttype in ('cr', 'credit', 'in', 'inflow', 'deposit', 'received', 'income'):
                    credit = amt
                    debit = Decimal(0)
                else:
                    # Use sign if available
                    if amt < 0:
                        debit = -amt
                        credit = Decimal(0)
                    else:
                        credit = amt
                        debit = Decimal(0)

            # Normalize decimals
            try:
                debit = Decimal(str(debit).replace(",", "")) if debit not in (None, '') else Decimal(0)
                credit = Decimal(str(credit).replace(",", "")) if credit not in (None, '') else Decimal(0)
            except Exception:
                debit = Decimal(0)
                credit = Decimal(0)
            if debit == 0 and credit == 0:
                continue
            txn = Transaction(
                account=acc,
                date=dt if hasattr(dt, 'year') else None,
                description=str(desc),
                reference=str(ref)[:100],
                debit=debit,
                credit=credit,
                created_by=user,
            )
            if txn.date is None:
                self.stderr.write("Skipping row without valid date")
                continue
            txn.save()
            created += 1

        if created == 0:
            self.stdout.write("Imported 0 transactions. Use --preview to inspect headers or provide --default-account if the sheet lacks per-row account names.")
        else:
            self.stdout.write(self.style.SUCCESS(f"Imported {created} transactions."))
